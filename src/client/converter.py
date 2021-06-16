import win32gui,win32con,win32api,win32ui
# import re
import pyautogui
import re, traceback
import time
import datetime
import sys
import tempfile
import argparse
import os

from PIL import Image
import pytesseract
import cv2

url = 'http://192.168.10.20:8004/api'
x = 0
y = 0 
w = 0
h = 0 
vHeighScreen = 0
vBookingCreatePage = False
vCuurentBookingMode = ''
x_capture =0
y_capture = 0
w_capture = 0
h_capture = 0

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class WindowMgr:
    """Encapsulates some calls to the winapi for window management"""
    settingFile =''


    def __init__ (self):
        """Constructor"""
        self.hwnd = None

    def find_window(self,title):
        try:
            self.hwnd = win32gui.FindWindow(None, title)
            assert self.hwnd
            return self.hwnd
        except:
            pyautogui.alert(text='Not found program name ' + title + '\n' 
                            'Please open program before excute script', title='Unable to open program', button='OK')
            print ('Not found program')
            return None


    def set_onTop(self,hwnd):
        win32gui.SetForegroundWindow(hwnd)
        return win32gui.GetWindowRect(hwnd)



    def Maximize(self,hwnd):
        win32gui.ShowWindow(hwnd,win32con.SW_RESTORE)#, win32con.SW_MAXIMIZE

    def get_mouseXY(self):
        return win32gui.GetCursorPos()

    def set_mouseXY(self):
        import os.path
        import json
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Current Window X : %s  Y: %s' %(x,y))
        fname = 'setting.json'
        if os.path.isfile(fname) :
            dict = eval(open(fname).read())
            x1 = dict['x']
            y1 = dict['y']
            print ('Setting X : %s  Y: %s' %(x1,y1))
        win32api.SetCursorPos((x+x1,y+y1))
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, x+x1, y+y1, 0, 0)
        win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, x+x1, y+y1, 0, 0)
        print ('Current Mouse X %s' % self.get_mouseXY()[0])
        print ('Current Mouse Y %s' % self.get_mouseXY()[1])


    def saveFirstDataPos(self):
        x,y,w,h = win32gui.GetWindowRect(self.hwnd)
        print ('Window X : %s  Y: %s' %(x,y))
        x1,y1 = self.get_mouseXY()
        print ('Mouse X : %s  Y: %s' %(x1,y1))
        data={}
        data['x'] = x1-x
        data['y'] = y1-y
        # f = open("setting.json", "w")
        # self.settingFile
        f = open(settingFile, "w")
        f.write(str(data))

        f.close()

    def wait(self,seconds=1,message=None):
        """pause Windows for ? seconds and print
an optional message """
        win32api.Sleep(seconds*1000)
        if message is not None:
            win32api.keybd_event(message, 0,0,0)
            time.sleep(.05)
            win32api.keybd_event(message,0 ,win32con.KEYEVENTF_KEYUP ,0)

    def typer(self,stringIn=None):
        PyCWnd = win32ui.CreateWindowFromHandle(self.hwnd)
        for s in stringIn :
            if s == "\n":
                self.hwnd.SendMessage(win32con.WM_KEYDOWN, win32con.VK_RETURN, 0)
                self.hwnd.SendMessage(win32con.WM_KEYUP, win32con.VK_RETURN, 0)
            else:
                print ('Ord %s' % ord(s))
                PyCWnd.SendMessage(win32con.WM_CHAR, ord(s), 0)
        PyCWnd.UpdateWindow()

    def WindowExists(windowname):
        try:
            win32ui.FindWindow(None, windowname)

        except win32ui.error:
            return False
        else:
            return True



def main():
    import json

    try:
        ldir = tempfile.mkdtemp()
        parser = argparse.ArgumentParser()
        parser.add_argument('-d', '--directory', default=ldir)
        args = parser.parse_args()
        tmpDir = args.directory
        print (tmpDir)

        fname = 'setting.json'

        # regex = "Untitled - Notepad"
        # regex = "Microsoft Excel - Book1"
        regex = "Session A - [24 x 80]"
        state_left = win32api.GetKeyState(0x01)  # Left button down = 0 or 1. Button up = -127 or -128
        state_right = win32api.GetKeyState(0x02)  # Right button down = 0 or 1. Button up = -127 or -128

        import os.path
        secs_between_keys=0.01

        win = WindowMgr()
        bCols =bcolors()
        win_ins = win.find_window(regex)
        # from colorama import Fore, Back, Style
        # print(Fore.RED + 'some red text')
        # print(Back.GREEN + 'and with a green background')
        # print(Style.DIM + 'and in dim text')
        # print(Style.RESET_ALL)
        # print('back to normal now')
        import sys
        # from colorama import Fore, Back, Style
        from colorama import init, AnsiToWin32,Fore, Back, Style
        init(wrap=False)
        stream = AnsiToWin32(sys.stderr).stream

        if win_ins == None :
            print (Fore.RED + ('Not found session : %s' % regex ), file=stream)
            sys.exit()
        else:
            print (Fore.GREEN + ('Start working on -- session : %s' % regex ), file=stream)



            #1) Activate Session ,put in onTop
            # print(pyautogui.size())
            win.Maximize(win_ins)
            global vHeighScreen
            global x
            global y
            global w
            global h
            global x_capture
            global y_capture
            global w_capture
            global h_capture

            (x,y,w,h) =win.set_onTop(win_ins)
            vHeighScreen = h*0.11

            if os.path.isfile(fname) :
                print ('Found Setting file')
                dict = eval(open(fname).read())
                x2 = x+dict['x']
                y2 = y+dict['y']
                w2 = dict['w'] - dict['x']
                h2 = dict['h']
            else:
                print ('Not found Setting file')
                x2 = x
                y2 = h-vHeighScreen-5
                w2 = w-x
                h2 = vHeighScreen/2

            x_capture = x2
            y_capture = y2
            w_capture = w2
            h_capture = h2


            # filename="images/test.png"
            
            # newTop = (y+h)-100
            # print (x,y,w,h,newTop)
            # pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files (x86)/Tesseract-OCR/tesseract'
            # # im = pyautogui.screenshot(filename,region=(x,h-vHeighScreen-5, w-x,vHeighScreen/3))
            # im = pyautogui.screenshot(filename,region=(x_capture,y_capture, w_capture,h_capture))
            # text = pytesseract.image_to_string(Image.open(filename))
            # print (text)



            # im = pyautogui.screenshot('sample.png',region=(x,h-vHeighScreen, w-x,vHeighScreen/3))
            
            #  Initial -Setup
            pyautogui.press('f12')
            pyautogui.press('f12')
            pyautogui.press('f12')
            pyautogui.press('f12')
            #2) Now On "Product Environment LCB1" screen.
            #Need to input "1" --> Work with CTCS.
            pyautogui.typewrite('1', interval=secs_between_keys)
            pyautogui.press('enter')

            #3) Now On "Select one of following" screen.
            #Need to input "1" --> Order.
            pyautogui.typewrite('1', interval=secs_between_keys)
            pyautogui.press('enter')

            #3) Now On "CTS order" screen.
            #Need to input "1" -->  Booking (EMPTY OUT / FULL IN).
            pyautogui.typewrite('1', interval=secs_between_keys)
            pyautogui.press('enter')

            #4)Goto Booking Order Creation
            pyautogui.press('f6')

            

            import glob
            file_list = glob.glob('files\*.xlsx')

            # print ('==============================================')
            print (Fore.GREEN + ('==============================================' ), file=stream)

            if len(file_list) == 0 :
                print (Fore.RED +'Not found any Excel file', file=stream)
                sys.exit()


            #Start Looping############
            from openpyxl import load_workbook
            import io
                            #Predefine data.....
            prev_booking=''
            curr_booking=''

            vLastBookExist = False
            vNextBooking = False
            vMode ='FIRST'
            global vBookingCreatePage
            vBookingCreatePage= True
            pterm='Y'
            vContainerCreateSuccess = False
            vBookingMode =''

            #import xlrd

            for shore_file in file_list:
                print (Fore.GREEN + ('File name : %s ' % shore_file ), file=stream)        
                # print ('File name : %s ' % shore_file )
                
                #Start to fill Booking
                
                # wb2 = load_workbook(shore_file,read_only=True)
                
                with open(shore_file, "rb") as f:
                    in_mem_file = io.BytesIO(f.read())

                wb2 = load_workbook(in_mem_file, read_only=True,data_only=True)
                # wb2 = xlrd.open_workbook(file_contents=in_mem_file.getvalue())

                

                # ws = wb2['sheet']
                for ws in wb2:
                    # print('Sheet name : %s' % ws.title)
                    print (Fore.GREEN + ('Sheet name : %s ' % ws.title ), file=stream)

                print(Style.RESET_ALL)

                # Open Log files
                filelog = open(shore_file.replace('.xlsx','.txt'), "a")

                filelog.write('Start On : %s \n' % datetime.datetime.now() )



                # print (ws['k6'].value,ws['k7'].value,ws['k8'].value)
                # sys.exit()
                # #Predefine data.....
                # prev_booking=''
                # curr_booking=''

                # vLastBookExist = False
                # vNextBooking = False
                # vMode ='FIRST'
                # global vBookingCreatePage
                # vBookingCreatePage= True
                # pterm='Y'
                # vContainerCreateSuccess = False
                # vBookingMode =''


                filename="images/message.png"
                pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files (x86)/Tesseract-OCR/tesseract'

                # for index,row in enumerate(ws.iter_rows()) :
                # for row in ws.rows :
                maxRow = 100
                print(Style.RESET_ALL)

                for index,row in enumerate(ws.iter_rows()) :
                    if index == maxRow :
                        break

                    fedder = row[6].value #Feeder -- Vessel
                    voy = row[7].value #voy
                    container_type = row[4].value #Container type
                    container_high = row[3].value.__str__()
                    container_long = row[2].value.__str__()
                    container = row[5].value #Container Number
                    shipper = row[0].value #shiper
                    booking =row[8].value #Booking Number
                    next_booking = ws.cell(row=index + 2, column=9).value
                    pod =row[10].value#POD
                    payment = row[22].value#Payment
                    dg_class = row[14].value #DG class
                    unno = row[15].value# unno number
                    line = row[1].value


                    # print (container)
                    # continue


                    ########################################
                    curr_booking = booking
                    ########################################

                    if dg_class==None:
                        dg_class=''

                    unno = row[15].value #unno number
                    if unno==None:
                        unno=''

                    temperature = row[12].value #temperature
                    if temperature==None:
                        temperature=0

                    #Accept only Container exist in File
                    if not (container != ' ' and container != '' and  container != None and booking != 'BK Number')   :
                        continue

                    #Check Container format
                    import re
                    rex = re.compile("^[A-Z]{4}[0-9]{7}$")
                    if not rex.match(container):
                        continue


                    # print (container)
                    # continue
                    
                    #Check Container Existing????
                    
                    chk_container = get_container('container',container.strip(),booking.strip())
                    if len(chk_container) > 0 :
                        print (Fore.YELLOW + ('####Container : %s : %s already exist!! ' % (container,booking)),file=stream)
                        print(Style.RESET_ALL)
                        filelog.write('Container : %s - %s already exist!!!!\n' % (container,booking))
                        continue





                    print (Fore.YELLOW + ('****Start process %s : %s ****' % (container,booking)),file=stream)
                    print(Style.RESET_ALL)
                    vessel_data = get_vessel('vessel',fedder.strip())
                    # print(vessel_data)
                    
                    if len(vessel_data) == 0 :
                        print (Fore.RED + ('Not found vessel %s on database' % fedder),file=stream)
                        print(Style.RESET_ALL)
                        sys.exit()

                    if  vessel_data[0]['code'] == '':
                        print (Fore.RED + ('Vessel code of %s is not configure' % fedder),file=stream)
                        print(Style.RESET_ALL)
                        sys.exit()

                    vessel_code = vessel_data[0]['code']


                    #Create Booking or Modify Booking
                    shipper_data = {'name':shipper.replace('+',' ').strip()}
                    shipper_id = create_shipper('shipper',shipper_data)
                    # print ('Shipper ID : %s' % shipper_id)

                    # Get or Create Vessel
                    vessel_data = v_name = {'name' : fedder.strip(),
                                        'code' : vessel_code.strip(),
                                        'description' : fedder.strip()}
                    vessel_id = create_vessel('vessel',vessel_data)
                    # print ('Vessel ID : %s' % vessel_id)

                    booking_data={ 'number' : booking.strip(),
                    'pod':pod.strip(),
                    'voy':voy.strip(),
                    'shipper': shipper_id,
                    'vessel':vessel_id,
                    'description':booking.strip()}

                    print(Style.RESET_ALL)
                    booking_id = create_booking('booking',booking_data)


                    if payment=='CASH':
                        pterm ='Y'
                    else:
                        pterm = 'N'
                    new_pod = pod[2:] + pod[:2]

                    if prev_booking != curr_booking :
                        global vCuurentBookingMode

                        vBookingCreated,vCuurentBookingMode = ctcs_create_booking(booking,line,line, \
                                vBookingMode,vContainerCreateSuccess)
                        vBookingMode = vCuurentBookingMode
                        prev_booking = booking
                        # sys.exit()
                    

                    vMsg=''
                    # print ('Process Container')
                    # print (vBookingCreated,container,shipper,vessel_code,voy,new_pod, \
                    #             pterm,container_long,container_high,container_type,booking_id,dg_class,unno,temperature)
                    vContainerCreateSuccess,vMsg = ctcs_create_container(vBookingCreated,container,shipper,vessel_code,voy,new_pod, \
                                pterm,container_long,container_high,container_type,booking_id,dg_class,unno,temperature)
                    vBookingCreated = False #Next container will be put only Conatainer

                    filelog.write('Create container : %s - %s - %s \n' % (booking,container,vMsg))
                    # sys.exit()


                filelog.write('Finished On : %s  -- Total %s containers \n' % (datetime.datetime.now(),index) )

                wb2._archive.close()
                wb2.close()
                # Delete file shore_file
                os.remove(shore_file)
                print ('##############Finished####################')



    except Exception as e:
    # except:
        if hasattr(e, 'message'):
            print(e.message)
        else:
            print(e)
        print('Error on line {}'.format(sys.exc_info()[-1].tb_lineno), type(e).__name__, e)
        f = open(tmpDir + "log.txt", "w")
        f.write(traceback.format_exc())


def ctcs_create_booking(booking,line,agent,vBookingMode,ContainerSuccess):
    filename="images/booking.png"
    pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files (x86)/Tesseract-OCR/tesseract'
    secs_between_keys=0.01
    global vBookingCreatePage
    ################Goto Create Order Page########################
    if not vBookingCreatePage :
        print ('Move to next Booking')
        print ('vBookingMode %s' % vBookingMode)
        # sys.exit()
        if vBookingMode == 'CHANGE':
            pyautogui.press('f12')
            pyautogui.press('f12')
            pyautogui.press('f12')
            pyautogui.press('f6')
            vBookingCreatePage=True
            # sys.exit()

        if vBookingMode == 'ADD':
            #Mode Add Booking ,No need for F6
            vBookingCreatePage = True
            pyautogui.press('f12')
            if ContainerSuccess :
                pyautogui.press('f12')
                print ('Create successful STOP')
            # sys.exit()
            # pyautogui.press('f6')
    ########################################
  
    pyautogui.typewrite(booking, interval=secs_between_keys)
    pyautogui.press('tab')
    pyautogui.typewrite(line, interval=secs_between_keys)
    pyautogui.press('tab')
    pyautogui.typewrite(agent, interval=secs_between_keys)
    pyautogui.press('enter')

    im = pyautogui.screenshot(filename,region=(x_capture,y_capture, w_capture,h_capture))
    text = pytesseract.image_to_string(Image.open(filename))
    if len(text)>0 :
        secs_between_keys = 0.05
        print ('Modify booking %s' % booking)
        created = False
        mode = "CHANGE"
        pyautogui.press('f12')
        pyautogui.typewrite(booking, interval=secs_between_keys)
        
        time.sleep(0.5)
        pyautogui.press('enter')
        pyautogui.typewrite('2', interval=secs_between_keys)
        pyautogui.press('enter')
        pyautogui.press('enter')
        pyautogui.press('f6') # Ready for Input Container
    else:
        created = True
        mode = "ADD"
        print ('Create booking %s' % booking)

    return created,mode


def ctcs_create_container(vContainerMode ,container,shipper,vessel_code, \
                        voy,pod,cash,lg,hg,ctype,booking_id,dg_class,unno,temperature):

    filename="images/container.png"
    pytesseract.pytesseract.tesseract_cmd = 'C:/Program Files (x86)/Tesseract-OCR/tesseract'
    secs_between_keys=0.01
    global vBookingCreatePage
    vBookingCreatePage = False 

    #Common data input
    pyautogui.press('down')
    pyautogui.typewrite(lg, interval=secs_between_keys) #Container Long
    pyautogui.press('tab')
    pyautogui.typewrite(hg, interval=secs_between_keys) #Container Height
    pyautogui.typewrite(ctype, interval=secs_between_keys) #Container Height
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('tab')
    pyautogui.press('down')

    if vContainerMode : #Add new mode
        pyautogui.typewrite(shipper[:34], interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.press('down')
        pyautogui.typewrite(container, interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.typewrite(vessel_code, interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.typewrite(voy.strip(), interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.typewrite(pod, interval=secs_between_keys)
        # pyautogui.press('down')
        #Now on DG Class box
        # not (dg_class != '' or dg_class != 'None')
        if len(dg_class)>0 :
            pyautogui.typewrite(dg_class.__str__(), interval=secs_between_keys)
            pyautogui.press('down')
            pyautogui.press('left',len(dg_class.__str__()))
            pyautogui.typewrite(unno, interval=secs_between_keys)
            pyautogui.press('tab')
        else :
            pyautogui.press('down')
            pyautogui.press('down')

        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.typewrite('AUTO EDI', interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.typewrite(cash, interval=secs_between_keys)
    else : #Change Mode
        # pyautogui.press('down')
        # pyautogui.press('down')
        pyautogui.typewrite(shipper[:34], interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.press('down')
        pyautogui.typewrite(container, interval=secs_between_keys)
        #Now on DG Class box

        if len(dg_class)>0 :
            pyautogui.press('tab')
            pyautogui.press('down')
            pyautogui.press('down')
            pyautogui.typewrite(dg_class.__str__(), interval=secs_between_keys)
            pyautogui.press('down')
            pyautogui.press('left',len(dg_class.__str__()))
            pyautogui.typewrite(unno, interval=secs_between_keys)
            pyautogui.press('tab')
        else :
            pyautogui.press('tab')
            pyautogui.press('down')
            pyautogui.press('down')
            pyautogui.press('down')
            pyautogui.press('down')

        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.press('down')
        pyautogui.typewrite('AUTO EDI', interval=secs_between_keys)
        pyautogui.press('tab')
        pyautogui.typewrite(cash, interval=secs_between_keys)


    # sys.exit()
    pyautogui.press('enter')

    im = pyautogui.screenshot(filename,region=(x_capture,y_capture, w_capture,h_capture))
    text = pytesseract.image_to_string(Image.open(filename))

    if len(text) > 5 :
        print ('Unable to create container %s : %s' % (container,text) )
        pyautogui.press('up')
        pyautogui.press('up')
        pyautogui.press('up')
        pyautogui.press('up')
        pyautogui.press('up')
        if not ('already in system' in text or 'Container is' in text or 'has already' in text or 'has ' in text or 'IN order:' in text):
            pyautogui.press('up')
        created = False
    else :
        print ('Create container %s successful!' % container )
        text ='Successful!!!'
        created = True


    container_data_dict = {
    'number': container,
    'booking': booking_id,
    'container_type': ctype,
    'container_size': lg,
    'description': container,
    'payment': cash,
    'dg_class': dg_class,
    'unno': unno,
    'temperature': temperature
    }

    # Save COntainer to DB
    c_id = create_container('container',container_data_dict)



    return created,text

  








def get_vessel(service,key_value):
    import urllib3
    import json
    http = urllib3.PoolManager()
    url_service = url + '/' + service + '/?name=' + key_value.replace(' ','%20')
    r = http.request('GET', url_service)
    if r.status == 200:
        str = r.data.decode("utf-8")
        data = json.loads(str)
    else :
        data={}
    # print (data)
    return data

def create_vessel(service,data):
    import urllib3
    import json
    http = urllib3.PoolManager()
    headers = {'Content-type': 'application/json'}
    # headers = urllib3.util.make_headers(basic_auth='admin:lcb12017',content_type='application/json')
    url_service = url + '/' + service + '/create/'

    new_data =get_shipper(service,data['name'])
    if len(new_data) == 0 :
        r = http.request('POST', url_service,headers=headers,body=json.dumps(data))
        new_data =get_vessel(service,data['name'])

    vessel_id = new_data[0]['id']

        # shipper_id=0
    return vessel_id

def get_shipper(service,key_value):
    import urllib3
    # from urllib import urlencode
    # import urllib
    import json
    http = urllib3.PoolManager()
    url_service = url + '/' + service + '/?name=' + key_value.replace(' ','%20')
    r = http.request('GET', url_service)
    if r.status == 200:
        str = r.data.decode("utf-8")
        data = json.loads(str)
    else :
        data={}
    # print (data)
    return data

def create_shipper(service,data):
    import urllib3
    import json
    http = urllib3.PoolManager()
    headers = {'Content-type': 'application/json'}
    # headers = urllib3.util.make_headers(basic_auth='admin:lcb12017',content_type='application/json')
    url_service = url + '/' + service + '/create/'

    new_data =get_shipper(service,data['name'])
    if len(new_data) == 0 :
        r = http.request('POST', url_service,headers=headers,body=json.dumps(data))
        new_data =get_shipper(service,data['name'])

    shipper_id = new_data[0]['id']

        # shipper_id=0
    return shipper_id

def get_booking(service,key_value):
    import urllib3
    import json
    http = urllib3.PoolManager()
    url_service = url + '/' + service + '/?number=' + key_value.replace(' ','%20')
    r = http.request('GET', url_service)
    if r.status == 200:
        str = r.data.decode("utf-8")
        data = json.loads(str)
    else :
        data={}
    # print (data)
    return data

def create_booking(service,data):
    import urllib3
    import json
    http = urllib3.PoolManager()
    headers = {'Content-type': 'application/json'}
    # headers = urllib3.util.make_headers(basic_auth='admin:lcb12017',content_type='application/json')
    url_service = url + '/' + service + '/create/'

    new_data =get_booking(service,data['number'])
    if len(new_data) == 0 :
        print ('Create New Booking in CRM')
        r = http.request('POST', url_service,headers=headers,body=json.dumps(data))
        new_data =get_booking(service,data['number'])

    booking_id = new_data[0]['id']

        # shipper_id=0
    return booking_id


def get_container(service,key_value,booking):
    import urllib3
    import json
    http = urllib3.PoolManager()
    url_service = url + '/' + service + '/?number=' + key_value.replace(' ','%20') +'&booking=' + ('%s' % booking)
    r = http.request('GET', url_service)
    if r.status == 200:
        str = r.data.decode("utf-8")
        data = json.loads(str)
    else :
        data={}
    # print (data)
    return data

def create_container(service,data):
    import urllib3
    import json
    http = urllib3.PoolManager()
    headers = {'Content-type': 'application/json'}
    # headers = urllib3.util.make_headers(basic_auth='admin:lcb12017',content_type='application/json')
    url_service = url + '/' + service + '/create/'

    new_data =get_container(service,data['number'],data['booking'])
    if len(new_data) == 0 :
        print ('Create New Container in CRM')
        r = http.request('POST', url_service,headers=headers,body=json.dumps(data))
        # new_data =get_container(service,data['number'],data['booking'])
        # container_id = new_data[0]['id']

    

        # shipper_id=0
    return ''



main()

#2558